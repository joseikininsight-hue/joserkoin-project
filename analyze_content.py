#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
助成金・補助金投稿のC列本文を分析し、SEO対策の問題点を洗い出す
"""

import openpyxl
from bs4 import BeautifulSoup
import re
from collections import Counter
import json

def analyze_html_structure(html_content):
    """HTML構造を分析"""
    if not html_content or not isinstance(html_content, str):
        return None
    
    soup = BeautifulSoup(html_content, 'html.parser')
    
    analysis = {
        'has_content': bool(html_content.strip()),
        'length': len(html_content),
        'has_html_tags': bool(soup.find()),
        'has_semantic_tags': False,
        'heading_structure': [],
        'css_classes': [],
        'inline_styles': [],
        'semantic_tags': [],
        'links_count': 0,
        'images_count': 0,
        'tables_count': 0,
        'lists_count': 0,
        'divs_count': 0,
        'sections_count': 0,
        'has_schema': False,
        'seo_issues': []
    }
    
    # 見出し構造の確認
    for level in range(1, 7):
        headings = soup.find_all(f'h{level}')
        if headings:
            analysis['heading_structure'].append({
                'level': level,
                'count': len(headings),
                'texts': [h.get_text()[:50] for h in headings[:3]]  # 最初の3つのみ
            })
    
    # セマンティックタグの確認
    semantic_tags = ['article', 'section', 'header', 'footer', 'nav', 'aside', 'main']
    for tag in semantic_tags:
        if soup.find(tag):
            analysis['semantic_tags'].append(tag)
            analysis['has_semantic_tags'] = True
    
    # CSSクラスの収集
    all_classes = []
    for tag in soup.find_all(class_=True):
        if isinstance(tag['class'], list):
            all_classes.extend(tag['class'])
        else:
            all_classes.append(tag['class'])
    analysis['css_classes'] = list(set(all_classes))[:20]  # 最初の20個
    
    # インラインスタイルの確認
    inline_styles = []
    for tag in soup.find_all(style=True):
        inline_styles.append(tag['style'][:100])  # 最初の100文字
    analysis['inline_styles'] = inline_styles[:5]  # 最初の5つ
    
    # その他の要素カウント
    analysis['links_count'] = len(soup.find_all('a'))
    analysis['images_count'] = len(soup.find_all('img'))
    analysis['tables_count'] = len(soup.find_all('table'))
    analysis['lists_count'] = len(soup.find_all(['ul', 'ol']))
    analysis['divs_count'] = len(soup.find_all('div'))
    analysis['sections_count'] = len(soup.find_all('section'))
    
    # Schema.orgの確認
    if soup.find(attrs={'itemtype': True}) or soup.find(attrs={'itemscope': True}):
        analysis['has_schema'] = True
    
    # SEO問題点の検出
    if not analysis['heading_structure']:
        analysis['seo_issues'].append('見出しタグ(H1-H6)が存在しない')
    elif not any(h['level'] == 2 for h in analysis['heading_structure']):
        analysis['seo_issues'].append('H2タグが存在しない（セクション構造が不明確）')
    
    if not analysis['has_semantic_tags']:
        analysis['seo_issues'].append('セマンティックHTML5タグが使用されていない')
    
    if not analysis['has_schema']:
        analysis['seo_issues'].append('Schema.org構造化データが存在しない')
    
    if analysis['inline_styles']:
        analysis['seo_issues'].append(f'インラインスタイルが{len(inline_styles)}箇所使用されている（CSS分離推奨）')
    
    if analysis['divs_count'] > 10 and not analysis['has_semantic_tags']:
        analysis['seo_issues'].append('divタグが多用されているがセマンティックタグが未使用')
    
    return analysis

def main():
    print("=" * 80)
    print("助成金・補助金投稿 C列本文 SEO分析レポート")
    print("=" * 80)
    print()
    
    # Excelファイルを開く
    wb = openpyxl.load_workbook('google_sheets_export.xlsx', read_only=True, data_only=True)
    ws = wb.active
    
    print(f"📊 シート名: {ws.title}")
    print(f"📊 総行数: {ws.max_row}")
    print()
    
    # ヘッダー行の確認
    headers = [cell.value for cell in ws[1]]
    print("📋 カラム構造:")
    for i, header in enumerate(headers[:10], 1):  # 最初の10列
        print(f"  {chr(64+i)}列: {header}")
    print()
    
    # C列（本文）を分析
    print("🔍 C列（本文）の詳細分析を開始...")
    print()
    
    all_analyses = []
    sample_contents = []
    
    # 2行目からデータ行
    max_row = ws.max_row if ws.max_row else 1000  # max_rowがNoneの場合は1000行まで
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=min(52, max_row)), start=2):
        post_id = row[0].value if row[0].value else f"Row{row_num}"
        title = row[1].value if len(row) > 1 and row[1].value else "タイトルなし"
        content = row[2].value if len(row) > 2 else None
        
        analysis = analyze_html_structure(content)
        if analysis:
            analysis['post_id'] = post_id
            analysis['title'] = title[:50]
            all_analyses.append(analysis)
            
            # 最初の3件をサンプルとして保存
            if len(sample_contents) < 3 and content:
                sample_contents.append({
                    'post_id': post_id,
                    'title': title,
                    'content_preview': content[:1000] if content else ""
                })
    
    print(f"✅ 分析完了: {len(all_analyses)}件の投稿を分析")
    print()
    
    # 統計情報
    print("=" * 80)
    print("📊 統計サマリー")
    print("=" * 80)
    print()
    
    total = len(all_analyses)
    has_html = sum(1 for a in all_analyses if a['has_html_tags'])
    has_semantic = sum(1 for a in all_analyses if a['has_semantic_tags'])
    has_schema = sum(1 for a in all_analyses if a['has_schema'])
    has_headings = sum(1 for a in all_analyses if a['heading_structure'])
    has_inline_styles = sum(1 for a in all_analyses if a['inline_styles'])
    
    print(f"総投稿数: {total}")
    print(f"HTMLタグあり: {has_html} ({has_html/total*100:.1f}%)")
    print(f"セマンティックHTML5使用: {has_semantic} ({has_semantic/total*100:.1f}%)")
    print(f"Schema.org使用: {has_schema} ({has_schema/total*100:.1f}%)")
    print(f"見出しタグ使用: {has_headings} ({has_headings/total*100:.1f}%)")
    print(f"インラインスタイル使用: {has_inline_styles} ({has_inline_styles/total*100:.1f}%)")
    print()
    
    # 見出し構造の統計
    print("=" * 80)
    print("📝 見出し構造の分析")
    print("=" * 80)
    print()
    
    heading_usage = Counter()
    for analysis in all_analyses:
        for h in analysis['heading_structure']:
            heading_usage[f"H{h['level']}"] += 1
    
    if heading_usage:
        print("見出しタグの使用頻度:")
        for tag, count in sorted(heading_usage.items()):
            print(f"  {tag}: {count}件 ({count/total*100:.1f}%)")
    else:
        print("⚠️ 見出しタグが全く使用されていません")
    print()
    
    # CSSクラスの統計
    print("=" * 80)
    print("🎨 CSSクラスの使用状況")
    print("=" * 80)
    print()
    
    all_classes = []
    for analysis in all_analyses:
        all_classes.extend(analysis['css_classes'])
    
    if all_classes:
        class_counter = Counter(all_classes)
        print(f"使用されているCSSクラス総数: {len(set(all_classes))}")
        print("\n頻出CSSクラス Top 20:")
        for cls, count in class_counter.most_common(20):
            print(f"  .{cls}: {count}回")
    else:
        print("⚠️ CSSクラスが使用されていません")
    print()
    
    # SEO問題点の集計
    print("=" * 80)
    print("⚠️ SEO問題点の集計")
    print("=" * 80)
    print()
    
    all_issues = []
    for analysis in all_analyses:
        all_issues.extend(analysis['seo_issues'])
    
    if all_issues:
        issue_counter = Counter(all_issues)
        print(f"検出された問題の総数: {len(all_issues)}")
        print("\n頻出問題 Top 10:")
        for issue, count in issue_counter.most_common(10):
            print(f"  • {issue}: {count}件 ({count/total*100:.1f}%)")
    else:
        print("✅ 重大なSEO問題は検出されませんでした")
    print()
    
    # サンプルコンテンツの表示
    print("=" * 80)
    print("📄 サンプルコンテンツ（最初の3件）")
    print("=" * 80)
    print()
    
    for i, sample in enumerate(sample_contents, 1):
        print(f"\n--- サンプル {i} ---")
        print(f"投稿ID: {sample['post_id']}")
        print(f"タイトル: {sample['title']}")
        print(f"\n本文プレビュー（最初の1000文字）:")
        print("-" * 80)
        print(sample['content_preview'])
        print("-" * 80)
        print()
    
    # 詳細分析結果をJSONで保存
    print("=" * 80)
    print("💾 詳細分析結果を保存中...")
    print("=" * 80)
    print()
    
    with open('content_analysis_report.json', 'w', encoding='utf-8') as f:
        json.dump({
            'summary': {
                'total_posts': total,
                'has_html_tags': has_html,
                'has_semantic_tags': has_semantic,
                'has_schema': has_schema,
                'has_headings': has_headings,
                'has_inline_styles': has_inline_styles
            },
            'heading_usage': dict(heading_usage),
            'common_classes': dict(class_counter.most_common(50)) if all_classes else {},
            'common_issues': dict(issue_counter.most_common(20)) if all_issues else {},
            'detailed_analyses': all_analyses[:20]  # 最初の20件の詳細
        }, f, ensure_ascii=False, indent=2)
    
    print("✅ 詳細レポートを content_analysis_report.json に保存しました")
    print()
    
    # 推奨される改善策
    print("=" * 80)
    print("💡 推奨される改善策")
    print("=" * 80)
    print()
    
    print("1. セマンティックHTML5構造の導入")
    print("   • <article> タグで投稿全体をラップ")
    print("   • <section> タグでコンテンツセクションを分割")
    print("   • <header> タグでヘッダー情報をマークアップ")
    print()
    
    print("2. 見出し階層の統一")
    print("   • H1: 投稿タイトル（1つのみ）")
    print("   • H2: メインセクション（対象者、内容、申請方法など）")
    print("   • H3: サブセクション")
    print()
    
    print("3. Schema.org構造化データの追加")
    print("   • GovernmentService または Grant スキーマの実装")
    print("   • JSON-LD形式での埋め込み")
    print()
    
    print("4. CSS設計の統一")
    print("   • BEM命名規則の採用（.grant-[block]__[element]--[modifier]）")
    print("   • インラインスタイルの排除")
    print("   • レスポンシブデザインの実装")
    print()
    
    print("5. アクセシビリティの向上")
    print("   • ARIA属性の適切な使用")
    print("   • alt属性の設定")
    print("   • フォーカス管理の改善")
    print()
    
    wb.close()
    print("=" * 80)
    print("分析完了！")
    print("=" * 80)

if __name__ == '__main__':
    main()
